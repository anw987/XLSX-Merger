import pandas as pd

# Input files
files = {
    "ATM_Status": "File1.xlsx",
    "ATM_STATUS_DETAIL": "File2.xlsx",
    "ATM_STATUS_WARNING": "File3.xlsx",
    "TRANSACTION": "File4.xlsx",
    "BASEMAN": "File5.xlsx"
}

# Output file
from datetime import datetime

timestamp = datetime.now().strftime("%y%m%d%H%M")
output_file = f"MergedFile-{timestamp}.xlsx"
#output_file = "merged.xlsx"

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, file_path in files.items():
        try:
            df = pd.read_excel(file_path)
            if sheet_name == "ATM_STATUS_WARNING":
               print("got it")
               df = df[["termid","Location","State","Status","category","Fault1","Fault2","Fault3","Fault4","Fault5","Fault6","Fault7","Fault8","admphone","city"]]
               df["Owner"] = "Owner Name"

               info_rows = pd.DataFrame([
                  ["Dashboard:", "/Dashboard/ATMWarning", "","","","","","","","","","","","","",""],
                  ["Name:", "ATM_Status_Warning", "", "","","","","","","","","","","",""],
                  ["Title:", "ATM STATUS WARNING","","","","","","","","","","","","",""],["","","","","","","","","","","","","","",""]
               ], columns=df.columns[:16])

               header_rows = pd.DataFrame([
                  ["ID", "Location", "State", "Line", "Desc", "Err 01", "Err 02", "Err 03", "Err 04", "Err 05", "Err 06", "Err 07", "Err 08", "Tel No", "City", "Ownr"]
               ], columns=df.columns[:16])

               df = pd.concat([info_rows, header_rows, df], ignore_index=True)

            if sheet_name == "BASEMAN":
               print("got it")
               df = df[["description","OBJECT_STATE"]]

               info_rows = pd.DataFrame([
                  ["Dashboard:", "/Dashboard/State"],
                  ["Name:", "BASEMAN"],
                  ["Title:", "STATUS KONEKSI H2H"],["",""]
               ], columns=df.columns[:2])

               header_rows = pd.DataFrame([
                  ["H2H", "STATUS"]
               ], columns=df.columns[:2])

               df = pd.concat([info_rows, header_rows, df], ignore_index=True)


            if sheet_name == "ATM_STATUS_DETAIL":
               print("got it")
               df = df[["termid","Location","State","Status","category","Fault1","Fault2","Fault3","Fault4","Fault5","Fault6","Fault7","Fault8","admphone","city"]]
               df["Owner"] = "Owner Name"

               info_rows = pd.DataFrame([
                  ["Dashboard:", "/Dashboard/ATMCritical", "","","","","","","","","","","","","",""],
                  ["Name:", "ATM_Status_Detail", "", "","","","","","","","","","","",""],
                  ["Title:", "ATM STATUS CRITICAL","","","","","","","","","","","","",""],["","","","","","","","","","","","","","",""]
               ], columns=df.columns[:16])

               header_rows = pd.DataFrame([
                  ["ID", "Location", "State", "Line", "Desc", "Err 01", "Err 02", "Err 03", "Err 04", "Err 05", "Err 06", "Err 07", "Err 08", "Tel No", "City", "Ownr"]
               ], columns=df.columns[:16])

               df = pd.concat([info_rows, header_rows, df], ignore_index=True)

            if sheet_name == "TRANSACTION":
               print("got it")
               info_rows = pd.DataFrame([
                  ["Dashboard:", "/Dashboard/Transaction"],
                  ["Name:", "TRANSACTION"],
                  ["Title:", "TRANSACTION SUMMARY"],["",""],["Series Name","Legend Content"]
               ], columns=df.columns[:2])

               df = pd.concat([info_rows], ignore_index=True)


            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            print(f"Processed: {file_path} -> {sheet_name}")


            from openpyxl.utils import get_column_letter

            worksheet = writer.sheets[sheet_name]

            for col_idx, col in enumerate(worksheet.columns, start=1):
               max_length = 0
               col_letter = get_column_letter(col_idx)

               for cell in col:
                  try:
                     if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                  except:
                     pass

               adjusted_width = max_length + 2  # padding
               worksheet.column_dimensions[col_letter].width = adjusted_width

            from openpyxl.styles import PatternFill, Font

            ws = writer.sheets[sheet_name]


            header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # grey
            dark_fill   = PatternFill(start_color="404040", end_color="404040", fill_type="solid")  # dark grey
            light_fill  = PatternFill(start_color="383838", end_color="383838", fill_type="solid")  # light grey
            green  = PatternFill(start_color="00DD00", end_color="00DD00", fill_type="solid")  # green

            white_font = Font(color="FFFFFF", bold=True)

            if sheet_name != "ATM_Status":
               for cell in ws[5]:
                  cell.fill = header_fill
                  cell.font = white_font

               for row_idx in range(6, ws.max_row + 1):
                  fill = dark_fill if (row_idx % 2 == 0) else light_fill

                  for cell in ws[row_idx]:
                     cell.fill = fill
                     cell.font = Font(color="FFFFFF")

               if sheet_name == "BASEMAN":
                  for row_idx in range(6, ws.max_row + 1):
                     fill = green

                     for cell in ws[row_idx]:
                        cell.fill = fill
                        cell.font = Font(color="000000")

        except Exception as e:
            print(f"Failed: {file_path}")
            print(e)

    df_chart = pd.read_excel("File1.xlsx")
    df_chart["XValue"] = pd.to_datetime(df_chart["XValue"], format="%Y-%m-%d %H.%M.%S")

    # Pivot: Approval vs Decline as columns
    pivot_df = df_chart.pivot(
        index="XValue",
        columns="Metric Name",
        values="Metric Value"
    ).reset_index()

    sheet_name = "TRANSACTION"

#    pivot_df.to_excel(writer, sheet_name=sheet_name, index=False)

    from openpyxl.chart import BarChart, Reference

    ws = writer.sheets[sheet_name]
    xvalues = pivot_df["XValue"].dt.strftime("%H:%M").tolist()
    # Write starting from column C
    for i, val in enumerate(xvalues, start=3):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}5"] = val

    chart = BarChart()
    chart.type = "col"
    chart.title = "ATM Approval vs Decline"

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(cats)

    ws.add_chart(chart, "A7")
    chart.width = 37
    chart.height = 12

    sheet_name = "ATM_Status"
    ws = writer.sheets[sheet_name]
    ws["C6"] = ws["B1"].value
    ws["C7"] = ws["C1"].value
    ws["C1"] = ""
    ws["A1"] = "Dashboard:"
    ws["A2"] = "Name:"
    ws["A3"] = "Title:"
    ws["B1"] = "/Dashboard/ATMStatus"
    ws["B2"] = "ATM_Status"
    ws["B3"] = "ATM PROBLEM"
    ws["A5"] = "Series Name"
    ws["A6"] = "Critical"
    ws["A7"] = "Warning"
    ws["B5"] = "Legend Content"
    ws["B6"] = "Critical - " + str(ws["C6"].value)
    ws["B7"] = "Warning - " + str(ws["C7"].value)

    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.text import CharacterProperties
    from openpyxl.chart.text import RichText

    chart = PieChart()
    chart.title = "ATM PROBLEM"
    chart.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=1400)  # ~14pt

    data = Reference(ws, min_col=3, min_row=6, max_row=7)
    labels = Reference(ws, min_col=1, min_row=6, max_row=7)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    pt1 = DataPoint(idx=0)
    pt1.graphicalProperties.solidFill = "5B9BD5"
    pt2 = DataPoint(idx=1)
    pt2.graphicalProperties.solidFill = "1F4E79"
    chart.series[0].data_points = [pt1, pt2]

    # Size (optional)
    chart.width = 35
    chart.height = 12

    # Position (upper center-ish)
    ws.add_chart(chart, "A9")

print("Merge completed: merged.xlsx")